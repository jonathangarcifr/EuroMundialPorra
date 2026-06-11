from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from apuestas.models import Apuesta
from apuestas.views import sincronizar_goleadores_partidos


class Command(BaseCommand):
    help = "Importa apuestas desde un Excel"

    def add_arguments(self, parser):
        parser.add_argument("ruta_excel", type=str)

    def handle(self, *args, **options):
        ruta_excel = options["ruta_excel"]

        wb = load_workbook(ruta_excel, data_only=True)
        ws = wb.active

        creadas = 0
        actualizadas = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            (
                nombre,
                email,
                pagado,
                equipo_1,
                equipo_2,
                equipo_3,
                equipo_4,
                equipo_5,
                equipo_6,
                equipo_7,
                equipo_8,
                equipo_9,
                equipo_10,
                equipo_11,
                equipo_12,
                goleador,
                equipo_goleador,
            ) = row

            if not nombre:
                continue

            datos = {
                "pagado": str(pagado).strip().lower() in ["sí", "si", "s", "true", "1"],
                "email": email,
                "equipo_1": equipo_1,
                "equipo_2": equipo_2,
                "equipo_3": equipo_3,
                "equipo_4": equipo_4,
                "equipo_5": equipo_5,
                "equipo_6": equipo_6,
                "equipo_7": equipo_7,
                "equipo_8": equipo_8,
                "equipo_9": equipo_9,
                "equipo_10": equipo_10,
                "equipo_11": equipo_11,
                "equipo_12": equipo_12,
                "goleador": goleador,
                "equipo_goleador": equipo_goleador,
            }

            _, creada = Apuesta.objects.update_or_create(
                nombre=nombre,
                defaults=datos,
            )

            if creada:
                creadas += 1
            else:
                actualizadas += 1

        sincronizar_goleadores_partidos()

        self.stdout.write(
            self.style.SUCCESS(
                f"Importación completada. Creadas: {creadas}. Actualizadas: {actualizadas}."
            )
        )